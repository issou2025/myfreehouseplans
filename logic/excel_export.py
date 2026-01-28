"""CivilQuant Pro — Export Excel (OpenPyXL).

Objectif:
- Générer un rapport Excel "pro" en mémoire (BytesIO), sans écrire sur disque.

Entrées:
- rows: liste de dicts (Désignation, Quantité, Unité, Catégorie)
- meta: informations complémentaires (nom fichier, échelle, date, etc.)
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def build_takeoff_excel_bytes(
    rows: list[dict[str, Any]],
    *,
    meta: dict[str, Any] | None = None,
    title: str = 'CivilQuant Pro — Métré DXF',
) -> BytesIO:
    """Construit un fichier Excel en mémoire."""

    meta = dict(meta or {})

    wb = Workbook()
    ws = wb.active
    ws.title = 'Métré'

    headers = ['Désignation', 'Quantité', 'Unité', 'Catégorie']

    # Titre
    ws['A1'] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    # En-têtes
    header_row = 3
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F2937')  # gris foncé
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Données
    start_data_row = header_row + 1
    for i, r in enumerate(rows):
        row_idx = start_data_row + i
        ws.cell(row=row_idx, column=1, value=str(r.get('Désignation', '') or ''))

        qty = r.get('Quantité', None)
        qty_cell = ws.cell(row=row_idx, column=2, value=qty)

        unit = str(r.get('Unité', '') or '')
        ws.cell(row=row_idx, column=3, value=unit)
        ws.cell(row=row_idx, column=4, value=str(r.get('Catégorie', '') or ''))

        # Formats numériques selon unité
        if isinstance(qty, (int, float)):
            if unit == 'm':
                qty_cell.number_format = '0.000'
            elif unit in {'m²', 'm2'}:
                qty_cell.number_format = '0.00'
            else:
                # unités / divers
                qty_cell.number_format = '0'

    last_row = start_data_row + max(len(rows) - 1, 0)

    # Style tableau
    if rows:
        table_ref = f"A{header_row}:D{last_row}"
        table = Table(displayName='TakeoffTable', ref=table_ref)
        style = TableStyleInfo(
            name='TableStyleMedium9',
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    # Largeurs colonnes
    col_widths = {
        1: 55,
        2: 14,
        3: 10,
        4: 18,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Bordures et alignements
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=header_row, max_row=max(last_row, header_row), min_col=1, max_col=4):
        for cell in row:
            cell.border = border
            if cell.row == header_row:
                continue
            if cell.column == 2:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.freeze_panes = f"A{start_data_row}"

    # Feuille meta
    meta_ws = wb.create_sheet('Meta')
    meta_ws['A1'] = 'Généré le'
    meta_ws['B1'] = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%SZ')

    row = 3
    for k in sorted(meta.keys()):
        meta_ws.cell(row=row, column=1, value=str(k))
        meta_ws.cell(row=row, column=2, value=str(meta.get(k)))
        row += 1

    meta_ws.column_dimensions['A'].width = 28
    meta_ws.column_dimensions['B'].width = 60

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
